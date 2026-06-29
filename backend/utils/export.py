"""
utils/export.py
Excel and CSV export helpers.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

import pandas as pd


def export_clients_excel(clients: list[dict]) -> bytes:
    """Export client list to Excel (.xlsx). Returns raw bytes."""
    df = pd.DataFrame(clients)
    _rename = {
        "code_client": "Code Client", "nom": "Nom", "email": "Email",
        "gouvernorat": "Gouvernorat", "nature_client": "Nature",
        "statut": "Statut", "score_actuel": "Score Risque",
        "risk_level": "Niveau Risque", "plafond_credit": "Plafond Crédit",
        "credit_utilise": "Crédit Utilisé", "total_impaye": "Total Impayé",
        "derniere_analyse": "Dernière Analyse",
    }
    df = df.rename(columns=_rename)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Clients")
        ws = writer.sheets["Clients"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # Colour-code risk level
        risk_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Niveau Risque":
                risk_col = i
                break
        if risk_col:
            risk_colors = {
                "FAIBLE": "C6EFCE", "MOYEN": "FFEB9C",
                "ÉLEVÉ": "FFC7CE", "INCONNU": "EFEFEF",
            }
            for row in ws.iter_rows(min_row=2, min_col=risk_col, max_col=risk_col):
                for cell in row:
                    color = risk_colors.get(str(cell.value or ""), "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    buf.seek(0)
    return buf.read()


def export_predictions_excel(predictions: list[dict]) -> bytes:
    df  = pd.DataFrame(predictions)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Prédictions")
    buf.seek(0)
    return buf.read()


def export_clients_csv(clients: list[dict]) -> str:
    if not clients:
        return ""
    buf    = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=clients[0].keys())
    writer.writeheader()
    writer.writerows(clients)
    return buf.getvalue()


def export_bulk_results_excel(results: list[dict]) -> bytes:
    rows = []
    for r in results:
        rows.append({
            "Code Client":   r.get("code_client", ""),
            "Statut":        r.get("label", r.get("error", "ERREUR")),
            "Score Risque":  r.get("risk_score", ""),
            "Niveau":        r.get("risk_level", ""),
            "Probabilité %": r.get("probability_risk", ""),
            "Résumé IA":     r.get("ai_summary", ""),
            "Erreur":        r.get("error", ""),
        })
    df  = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Résultats Bulk")
    buf.seek(0)
    return buf.read()
